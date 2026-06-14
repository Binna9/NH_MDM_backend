import io
import logging
import uuid
import zipfile
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

from app.models.nh_member_info import NhMemberInfo
from app.schemas.member import MemberExcelUploadError, MemberExcelUploadResponse, MemberExcelValidateResponse

TEMPLATE_FILENAME = "nh_member_upload_template.xlsx"
EXPORT_FILENAME = "nh_member_export.xlsx"
HEADER_ROW = 1
DATA_START_ROW = 2

EXCEL_HEADERS = ("순번", "성명", "실명번호", "고객번호", "핸드폰")
REQUIRED_DATA_HEADERS = ("성명", "실명번호", "고객번호", "핸드폰")
HEADER_TO_FIELD = {
    "성명": "nh_member_name",
    "실명번호": "nh_member_ssn",
    "고객번호": "nh_customer_no",
    "핸드폰": "nh_member_phone",
}

COLUMN_WIDTHS = (8, 14, 16, 14, 16)

_HEADER_FILL = PatternFill("solid", fgColor="1F6B2E")
_ROW_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")
_ROW_FILL_EVEN = PatternFill("solid", fgColor="EDF7EE")
_HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
_DATA_FONT = Font(name="맑은 고딕", size=10)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_THIN = Side(style="thin", color="BDBDBD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_sheet(sheet: Worksheet, data_row_count: int = 0) -> None:
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = width

    sheet.row_dimensions[HEADER_ROW].height = 22
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        cell = sheet.cell(row=HEADER_ROW, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    for row_idx in range(DATA_START_ROW, DATA_START_ROW + data_row_count):
        fill = _ROW_FILL_EVEN if row_idx % 2 == 0 else _ROW_FILL_ODD
        sheet.row_dimensions[row_idx].height = 18
        for col_idx in range(1, len(EXCEL_HEADERS) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = _DATA_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER


@dataclass
class ParsedMemberRow:
    row_number: int
    nh_member_name: str
    nh_member_ssn: str
    nh_customer_no: str
    nh_member_phone: str


def build_member_excel_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "조합원 업로드"

    for column_index, header in enumerate(EXCEL_HEADERS, start=1):
        sheet.cell(row=HEADER_ROW, column=column_index, value=header)

    _style_sheet(sheet, data_row_count=0)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_member_excel_export(
    db: Session,
    *,
    nh_member_name: str | None = None,
    nh_member_ssn: str | None = None,
    nh_customer_no: str | None = None,
    nh_member_phone: str | None = None,
    is_active: str | None = None,
) -> bytes:
    query = select(NhMemberInfo)

    if nh_member_name is not None:
        query = query.where(NhMemberInfo.nh_member_name.ilike(f"%{nh_member_name}%"))
    if nh_member_ssn is not None:
        query = query.where(NhMemberInfo.nh_member_ssn == nh_member_ssn)
    if nh_customer_no is not None:
        query = query.where(NhMemberInfo.nh_customer_no == nh_customer_no)
    if nh_member_phone is not None:
        query = query.where(NhMemberInfo.nh_member_phone == nh_member_phone)
    if is_active is not None:
        query = query.where(NhMemberInfo.is_active == is_active)

    members = db.scalars(
        query.order_by(NhMemberInfo.nh_member_name, NhMemberInfo.nh_customer_no)
    ).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "조합원 목록"

    for column_index, header in enumerate(EXCEL_HEADERS, start=1):
        sheet.cell(row=HEADER_ROW, column=column_index, value=header)

    for sequence, member in enumerate(members, start=1):
        sheet.append(
            [
                sequence,
                member.nh_member_name,
                member.nh_member_ssn,
                member.nh_customer_no,
                member.nh_member_phone,
            ]
        )

    _style_sheet(sheet, data_row_count=len(members))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _normalize_cell(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _parse_header_map(sheet: Worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for column_index in range(1, sheet.max_column + 1):
        header = _normalize_cell(sheet.cell(row=HEADER_ROW, column=column_index).value)
        if header in HEADER_TO_FIELD:
            header_map[header] = column_index

    missing_headers = [header for header in REQUIRED_DATA_HEADERS if header not in header_map]
    if missing_headers:
        raise ValueError(f"필수 헤더가 없습니다: {', '.join(missing_headers)}")

    return header_map


def _parse_member_rows(workbook_bytes: bytes) -> list[ParsedMemberRow]:
    try:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    except zipfile.BadZipFile:
        raise ValueError("올바른 엑셀 파일(.xlsx)이 아닙니다. 파일을 확인해 주세요.")
    except Exception as exc:
        raise ValueError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    try:
        sheet = workbook.active
        header_map = _parse_header_map(sheet)
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"엑셀 헤더를 읽는 중 오류가 발생했습니다: {exc}") from exc

    parsed_rows: list[ParsedMemberRow] = []
    for row_number in range(DATA_START_ROW, sheet.max_row + 1):
        values = {
            header: _normalize_cell(sheet.cell(row=row_number, column=column_index).value)
            for header, column_index in header_map.items()
        }

        if not any(values.get(header) for header in REQUIRED_DATA_HEADERS):
            continue

        parsed_rows.append(
            ParsedMemberRow(
                row_number=row_number,
                nh_member_name=values["성명"] or "",
                nh_member_ssn=values["실명번호"] or "",
                nh_customer_no=values["고객번호"] or "",
                nh_member_phone=values["핸드폰"] or "",
            )
        )

    workbook.close()
    return parsed_rows


def _validate_row(row: ParsedMemberRow) -> str | None:
    if not row.nh_member_name:
        return "성명은 필수입니다."
    if len(row.nh_member_name) > 50:
        return "성명은 50자 이하여야 합니다."
    if not row.nh_member_ssn:
        return "실명번호는 필수입니다."
    if len(row.nh_member_ssn) > 8:
        return "실명번호는 8자 이하여야 합니다."
    if not row.nh_customer_no:
        return "고객번호는 필수입니다."
    if len(row.nh_customer_no) > 10:
        return "고객번호는 10자 이하여야 합니다."
    if not row.nh_member_phone:
        return "핸드폰은 필수입니다."
    if len(row.nh_member_phone) > 13:
        return "핸드폰은 13자 이하여야 합니다."
    return None


def upsert_member_row(db: Session, row: ParsedMemberRow) -> str:
    existing = db.scalar(
        select(NhMemberInfo).where(NhMemberInfo.nh_customer_no == row.nh_customer_no)
    )

    action = "updated" if existing else "inserted"

    if existing:
        db.delete(existing)
        db.flush()

    member = NhMemberInfo(
        nh_member_id=str(uuid.uuid4()),
        nh_member_name=row.nh_member_name,
        nh_member_ssn=row.nh_member_ssn,
        nh_customer_no=row.nh_customer_no,
        nh_member_phone=row.nh_member_phone,
        is_active="Y",
    )
    db.add(member)

    try:
        db.flush()
    except IntegrityError as exc:
        raise ValueError(f"실명번호 또는 고객번호가 다른 조합원과 중복됩니다. (고객번호: {row.nh_customer_no})") from exc
    except SQLAlchemyError as exc:
        logger.exception("DB 오류 발생 (행 %d)", row.row_number)
        raise ValueError(f"데이터베이스 오류가 발생했습니다: {exc}") from exc

    return action


def _check_in_file_duplicates(row: ParsedMemberRow, seen_ssn: dict[str, int], seen_customer_no: dict[str, int]) -> str | None:
    if row.nh_member_ssn in seen_ssn:
        return f"엑셀 내 실명번호가 {seen_ssn[row.nh_member_ssn]}행과 중복됩니다."
    if row.nh_customer_no in seen_customer_no:
        return f"엑셀 내 고객번호가 {seen_customer_no[row.nh_customer_no]}행과 중복됩니다."
    return None


def validate_member_excel(db: Session, workbook_bytes: bytes) -> MemberExcelValidateResponse:
    try:
        parsed_rows = _parse_member_rows(workbook_bytes)
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    if not parsed_rows:
        raise ValueError("검증할 데이터가 없습니다.")

    would_insert = 0
    would_update = 0
    errors: list[MemberExcelUploadError] = []
    seen_ssn: dict[str, int] = {}
    seen_customer_no: dict[str, int] = {}

    for row in parsed_rows:
        validation_error = _validate_row(row)
        if validation_error:
            errors.append(MemberExcelUploadError(row=row.row_number, message=validation_error))
            continue

        in_file_error = _check_in_file_duplicates(row, seen_ssn, seen_customer_no)
        if in_file_error:
            errors.append(MemberExcelUploadError(row=row.row_number, message=in_file_error))
            continue

        existing = db.scalar(
            select(NhMemberInfo).where(NhMemberInfo.nh_customer_no == row.nh_customer_no)
        )
        if existing:
            would_update += 1
        else:
            would_insert += 1

        seen_ssn[row.nh_member_ssn] = row.row_number
        seen_customer_no[row.nh_customer_no] = row.row_number

    return MemberExcelValidateResponse(
        total=len(parsed_rows),
        would_insert=would_insert,
        would_update=would_update,
        failed=len(errors),
        errors=errors,
    )


def upload_member_excel(db: Session, workbook_bytes: bytes) -> MemberExcelUploadResponse:
    try:
        parsed_rows = _parse_member_rows(workbook_bytes)
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    if not parsed_rows:
        raise ValueError("업로드할 데이터가 없습니다.")

    inserted = 0
    updated = 0
    errors: list[MemberExcelUploadError] = []

    for row in parsed_rows:
        validation_error = _validate_row(row)
        if validation_error:
            errors.append(MemberExcelUploadError(row=row.row_number, message=validation_error))
            continue

        try:
            with db.begin_nested():
                action = upsert_member_row(db, row)
        except ValueError as exc:
            errors.append(MemberExcelUploadError(row=row.row_number, message=str(exc)))
            continue

        if action == "inserted":
            inserted += 1
        else:
            updated += 1

    if inserted > 0 or updated > 0:
        db.commit()
    else:
        db.rollback()

    return MemberExcelUploadResponse(
        inserted=inserted,
        updated=updated,
        failed=len(errors),
        errors=errors,
    )
